"""XLSX parser with multi-sheet row-wise header context extraction."""
from __future__ import annotations

import importlib.util
from typing import Any

from build_your_own_rag.models import (
    DocumentMetadata,
    ExtractionStrategy,
    ParsedDocument,
    ParsedPage,
    SourceDocument,
    utc_now_iso,
)
from build_your_own_rag.utils.logger import get_logger

logger = get_logger(__name__)


class XlsxParseError(RuntimeError):
    """Raised when XLSX parsing fails."""


def parse_xlsx(
    source: SourceDocument, strategy: ExtractionStrategy | None = None
) -> ParsedDocument:
    """Parse an XLSX document across all sheets, prepending headers to each row."""
    # Use FAST as default if no strategy is provided
    strategy = strategy or ExtractionStrategy.FAST
    parser_name = "xlsx_native"
    processing_started_at = utc_now_iso()
    status = "success"

    if importlib.util.find_spec("openpyxl") is None:
        raise RuntimeError("openpyxl is not installed. Run `pip install openpyxl`.")
    
    import openpyxl

    row_blocks = []
    total_row_count = 0
    sheet_count = 0
    sheet_names = []
    
    try:
        # data_only=True ensures we get formula results rather than the formula string itself
        wb = openpyxl.load_workbook(str(source.path), data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        sheet_count = len(sheet_names)
        
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            
            headers = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue  # Skip empty rows
                
                if not headers:
                    # Treat the first non-empty row as headers
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                    continue
                
                total_row_count += 1
                
                cell_texts = []
                for j, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_val = str(cell).strip()
                    if not cell_val:
                        continue
                        
                    header = headers[j] if j < len(headers) else f"Column_{j+1}"
                    if not header:
                        header = f"Column_{j+1}"
                        
                    cell_texts.append(f"{header}: {cell_val}")
                
                if cell_texts:
                    # Prepend sheet name context to each row block
                    row_block = f"[Sheet: {sheet_name}] " + " | ".join(cell_texts)
                    row_blocks.append(row_block)
                    
        wb.close()
    except Exception as exc:
        status = "error"
        logger.error(
            "Failed to parse XLSX",
            exc_info=exc,
            extra={"extra_info": {"source_id": source.source_id, "path": str(source.path)}}
        )
        raise XlsxParseError(f"Failed to parse XLSX {source.path}: {exc}") from exc

    text = "\n\n".join(row_blocks)
    processing_finished_at = utc_now_iso()

    format_metadata = {
        "xlsx": {
            "sheet_count": sheet_count,
            "sheet_names": sheet_names,
            "total_data_rows": total_row_count,
        }
    }

    metadata = DocumentMetadata(
        core={
            "filename": source.filename,
            "extension": source.extension,
            "source_path": str(source.path),
            "size_bytes": source.size_bytes,
            "mime_type": source.mime_type,
            "source_id": source.source_id,
            "source_type": source.source_type,
        },
        processing={
            "parser_name": parser_name,
            "requested_parser": "xlsx_native",
            "strategy": strategy.value,
            "ocr_used": False,
            "processing_status": status,
            "started_at": processing_started_at,
            "finished_at": processing_finished_at,
            "docling_error": None,
        },
        format_specific=format_metadata,
    )

    return ParsedDocument(
        source=source,
        text=text,
        pages=[],
        metadata=metadata,
        parser_name=parser_name,
        strategy=strategy,
        ocr_used=False,
        status=status,
    )

def register() -> None:
    """Register the XLSX parser."""
    from build_your_own_rag.parsing.parser_registry import register_parser
    register_parser(".xlsx", parse_xlsx)
